import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import os

files = os.listdir("app/data")

columns_to_read = [2, 3, 6, 41]


for index, file in enumerate(files):
    print(f"{index}. {file}")

print("*** **** *** *** *** ***")
index = int(input("Input number of the file: "))
print(f"*** **** *** Starting reading {files[index]} *** *** ***")

df = pd.read_excel(f"data/{files[index]}", usecols=columns_to_read)

print("*** **** *** File is ready *** *** ***")


df_filtered = df[df['Unnamed: 2'].astype(str).str.startswith("Взаємодія")]
df_filtered = df_filtered.sort_values(by=['Unnamed: 41', 'Unnamed: 6'])
wb = Workbook()
ws = wb.active
ws.title = "Error Report"


ws.append(["Error_Code", "Employee", "UUID", "Type"])


for index, row in df_filtered.iterrows():
    type_value = row['Unnamed: 2']  # Type
    employee = row['Unnamed: 6']  # Employee
    uuid = row['Unnamed: 3']  # UUID
    error_code = row['Unnamed: 41']  # Error_Code

    ws.append([error_code, employee,  uuid, type_value])

color_map = {}
fill_colors = ["FFFF99", "FF9999", "99FF99", "9999FF", "FF66CC"]

# Коліруємо клітинки для різних кодів помилок
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
    error_code = row[0].value
    if error_code not in color_map:
        color_map[error_code] = fill_colors[len(color_map) % len(fill_colors)]
    fill = PatternFill(start_color=color_map[error_code], end_color=color_map[error_code], fill_type="solid")
    row[0].fill = fill


wb.save("reports/report.xlsx")
print("Звіт збережено як report.xlsx")